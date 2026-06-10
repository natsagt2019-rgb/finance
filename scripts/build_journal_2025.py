# -*- coding: utf-8 -*-
"""
Ерөнхий журнал (Тайлан TT 2025-12-31.xlsx) → journal_entries seed.

- Gbalance-ийн эхний үлдэгдлийг 2024-12-31-ний opening бичилт болгож оруулна
  (журнал бүрэн эх сурвалж болж, баланс журналаас гарна).
- 2025 оны бүх гүйлгээг journal_entries болгоно.
- journal_account_balances-аас гүйлгээ баланс/тайлан тулж байгааг шалгана.

Гаралт: scripts/journal-2025.sql
"""
import openpyxl

XLSX = r"C:/Natsag/2023 on/Tailan 2025/Тайлан TT 2025-12-31.xlsx"
OUT = r"C:/finance 2.0/scripts/journal-2025.sql"
YEAR = 2025


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").replace(" ", ""))
    except ValueError:
        return 0.0


def code_of(v):
    if isinstance(v, (int, float)):
        s = str(int(v))
    else:
        s = str(v or "").strip()
    return s if s.isdigit() and len(s) >= 4 else None


def acc_kind(code):
    if code[0] in "12":
        return "asset"
    if code[0] == "3":
        return "liability"
    if code[0] == "4":
        return "equity"
    if code[:3] in ("510", "520", "840", "850"):
        return "income"
    return "expense"


wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 1. Эхний үлдэгдэл (Gbalance opening C-D) → opening бичилт ─────────────
gb = wb["Gbalance"]
opening = {}  # code -> net (debit-positive)
for r in gb.iter_rows(min_row=1, values_only=True):
    code = code_of(r[0])
    if not code:
        continue
    odt = num(r[2]) if len(r) > 2 else 0
    okt = num(r[3]) if len(r) > 3 else 0
    net = odt - okt
    if abs(net) > 0.001:
        opening[code] = net

# ── 2. 2025 журналын гүйлгээ ─────────────────────────────────────────────
jr = wb["Ерөнхий журнал"]
entries = []  # (no, date, desc, p_code, p_name, amount, dt, kt, cf)
for r in jr.iter_rows(min_row=5, values_only=True):
    amt = num(r[5]) if len(r) > 5 else 0
    dt = code_of(r[6]) if len(r) > 6 else None
    kt = code_of(r[7]) if len(r) > 7 else None
    if dt is None and kt is None:
        continue
    no = int(r[0]) if isinstance(r[0], (int, float)) else None
    d = r[1]
    date_s = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else f"{YEAR}-01-01"
    desc = (str(r[2]).strip() if len(r) > 2 and r[2] else None)
    p_code = (str(r[3]).strip() if len(r) > 3 and r[3] else None)
    p_name = (str(r[4]).strip() if len(r) > 4 and r[4] else None)
    cf = (str(r[8]).strip() if len(r) > 8 and r[8] else None)
    if cf and not cf[0].isdigit():
        cf = None
    entries.append((no, date_s, desc, p_code, p_name, amt, dt, kt, cf))


def q(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(round(v, 2)) if isinstance(v, float) else str(v)
    return "'" + str(v).replace("'", "''") + "'"


cols = ["entry_no", "txn_date", "description", "partner_code", "partner_name",
        "amount", "debit_code", "credit_code", "cf_code", "is_opening"]

vals = []
# Opening бичилт (нэг талт: дансаа байгаа тал руу, 2024-12-31)
for code, net in sorted(opening.items()):
    if net > 0:
        rec = [None, "2024-12-31", "Эхний үлдэгдэл", None, None, net, code, None, None, True]
    else:
        rec = [None, "2024-12-31", "Эхний үлдэгдэл", None, None, -net, None, code, None, True]
    vals.append("  (" + ", ".join(q(x) for x in rec) + ")")
# 2025 гүйлгээ
for (no, date_s, desc, pc, pn, amt, dt, kt, cf) in entries:
    rec = [no, date_s, desc, pc, pn, amt, dt, kt, cf, False]
    vals.append("  (" + ", ".join(q(x) for x in rec) + ")")

sql = (
    "-- " + "=" * 58 + "\n"
    "-- journal_entries seed — Түмэн Тээх ерөнхий журнал\n"
    f"-- Эхний үлдэгдэл {len(opening)} бичилт (2024-12-31) + 2025 гүйлгээ {len(entries)}.\n"
    "-- accounts-seed.sql-ийн ДАРАА ажиллуулна.\n"
    "-- " + "=" * 58 + "\n\n"
    "DELETE FROM journal_entries;\n\n"
    f"INSERT INTO journal_entries ({', '.join(cols)}) VALUES\n"
    + ",\n".join(vals) + ";\n"
)
open(OUT, "w", encoding="utf-8").write(sql)

# ── 3. ШАЛГАЛТ: журналаас тооцоод тулгана ────────────────────────────────
from collections import defaultdict
net_all = defaultdict(float)   # бүх бичилт (closing)
net_open = defaultdict(float)  # opening бичилт
net_2025 = defaultdict(float)  # 2025 турновер
cf_sum = defaultdict(float)
for code, n in opening.items():
    net_all[code] += n
    net_open[code] += n
for (no, date_s, desc, pc, pn, amt, dt, kt, cf) in entries:
    if dt:
        net_all[dt] += amt
        net_2025[dt] += amt
    if kt:
        net_all[kt] -= amt
        net_2025[kt] -= amt
    if cf:
        cf_sum[cf] += amt

clDt = sum(v for v in net_all.values() if v > 0)
clKt = -sum(v for v in net_all.values() if v < 0)
assets = sum(v for c, v in net_all.items() if acc_kind(c) == "asset")
liabEq = -sum(v for c, v in net_all.items()
              if acc_kind(c) in ("liability", "equity"))
net_profit = -sum(net_2025[c] for c in net_2025
                  if acc_kind(c) in ("income", "expense") and c[:3] != "920")
cf_net = sum((-v if ".2." in k else v) for k, v in cf_sum.items())

f = lambda x: f"{round(x):,}"
print("written", OUT)
print(f"Журнал: {len(entries)} гүйлгээ + {len(opening)} opening бичилт")
print(f"Гүйлгээ баланс Дт=Кт:  {f(clDt)} = {f(clKt)}  {'OK' if abs(clDt-clKt)<1 else 'ZORUU'}")
print(f"Баланс Актив=Өр+Өмч:   {f(assets)} = {f(liabEq)}  зөрүү={assets-liabEq:.2f}  (хүл 1,173,924,009)")
print(f"Орлого Цэвэр ашиг:     {f(net_profit)}  (хүл 469,341,122)")
print(f"Мөнгөн гүйлгээ цэвэр:  {f(cf_net)}  (хүл 147,854,179)")
