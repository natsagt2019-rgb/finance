# -*- coding: utf-8 -*-
"""
Түмэн Тээх ХХК-ийн бодит дансны төлөвлөгөө (Дансны_Төлөвлөгөө_TT.xlsx) -> Supabase seed.

Бүтэц: код AABBCC (AA=ангилал/класс, BB=бүлэг, CC=нарийвчилсан).
- Excel-д зөвхөн навч (leaf) дансууд кодтой (151). Класс/бүлгийн толгойнууд кодгүй.
- Энд класс (AA0000) ба бүлэг (AABB00) эх дансыг автоматаар үүсгэж модлог бүтэц гаргана.
- Хоёр хэл (name/name_en), тайлбар (note) хадгална.
- 'Зардлын Задаргаа' хуудаснаас зардлын дансны МУСГАА/татвар мэдээллийг note-д нэмнэ.

Гаралт: scripts/accounts-seed.sql
"""
import openpyxl

XLSX = r"C:/Users/natsa/Desktop/2026/Дансны_Төлөвлөгөө_TT.xlsx"
OUT = r"C:/finance 2.0/scripts/accounts-seed.sql"

ANGILAL_TYPE = {
    "Хөрөнгө": "asset",
    "Өр төлбөр": "liability",
    "Өмч": "equity",
    "Орлого": "income",
    "ББӨ": "expense",
    "Үйл ажиллагааны зардал": "expense",
    "Татвар": "expense",
}
DEBIT_NORMAL = {"asset", "expense"}  # -> nature Актив, бусад -> Пассив


def nature_of(t):
    return "Актив" if t in DEBIT_NORMAL else "Пассив"


def split_en(name):
    """'Мөнгө ба банкны данс (Cash & Bank)' -> ('Мөнгө ба банкны данс', 'Cash & Bank')"""
    if name and name.endswith(")") and "(" in name:
        i = name.rfind("(")
        return name[:i].strip(), name[i + 1 : -1].strip()
    return name, None


wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 1. Класс нэрс (AA0000) — Хураангуй хуудаснаас ────────────────────────
class_names = {}  # 'AA' -> (mon, en)
for r in wb["Хураангуй"].iter_rows(values_only=True):
    code = str(r[0]) if r[0] else ""
    if code[:2].isdigit() and "x" in code:
        mon, en = split_en(r[1])
        class_names[code[:2]] = (mon, en)

# ── 2. Зардлын задаргаа -> note нэмэлт ───────────────────────────────────
exp_meta = {}  # code -> note мөр
for r in wb["Зардлын Задаргаа"].iter_rows(min_row=2, values_only=True):
    if not r[0]:
        continue
    code = str(r[0]).strip()
    parts = []
    if r[3]:
        parts.append(str(r[3]).strip())  # ББӨ / Үйл ажиллагааны / ...
    if r[4]:
        parts.append(str(r[4]).strip())  # МУСГАА ангилал
    if r[5]:
        parts.append("татвар: " + str(r[5]).strip())  # хасагдах эсэх
    if parts:
        exp_meta[code] = " · ".join(parts)

# ── 3. Үндсэн хуудсыг алхаж навч + бүлгийн нэр цуглуулах ──────────────────
ws = wb["Дансны Төлөвлөгөө"]
leaves = []          # dict-үүд
group_names = {}     # 'AABB' -> (mon, en)
pending = []         # сүүлийн навчнаас хойш гарсан кодгүй толгойнууд (mon,en)

for r in ws.iter_rows(min_row=2, values_only=True):
    code = str(r[0]).strip() if r[0] else ""
    mon = (r[1] or "").strip()
    en = (r[2] or "").strip() or None
    ang = (r[3] or "").strip()
    note = (r[5] or "").strip() or None

    if not code:
        # толгой мөр — Ангилалтай бол бүлэг/классын нэр болж болзошгүй
        if mon:
            pending.append((mon, en))
        continue

    t = ANGILAL_TYPE.get(ang, "expense")
    # BB=00 бол тусдаа бүлэггүй — навч шууд класс дор (код давхцал гарахгүй)
    aabb = code[:4] if code[2:4] != "00" else None
    # энэ бүлгийн нэр = яг өмнөх толгой
    if aabb and aabb not in group_names and pending:
        group_names[aabb] = pending[-1]
    pending = []

    # зардлын мета note-д нэмэх
    extra = exp_meta.get(code)
    full_note = note
    if extra:
        full_note = (note + " | " + extra) if note else extra

    leaves.append({
        "code": code, "name": mon, "name_en": en, "type": t,
        "note": full_note,
    })

# ── 4. Класс ба бүлгийн эх дансуудыг үүсгэх ──────────────────────────────
def aabb_of(code):
    """BB=00 бол бүлэггүй (None) — навч шууд класс дор."""
    return code[:4] if code[2:4] != "00" else None

# Навчны type-аар класс/бүлгийн type-ийг тодорхойлно
group_type, class_type = {}, {}
for lf in leaves:
    aa, aabb = lf["code"][:2], aabb_of(lf["code"])
    class_type.setdefault(aa, lf["type"])
    if aabb:
        group_type.setdefault(aabb, lf["type"])

rows = []  # эцсийн дараалал: класс -> бүлэг -> навч (кодоор эрэмбэлж)
seen_class, seen_group = set(), set()
for lf in sorted(leaves, key=lambda x: x["code"]):
    aa, aabb = lf["code"][:2], aabb_of(lf["code"])
    if aa not in seen_class:
        seen_class.add(aa)
        cmon, cen = class_names.get(aa, (f"{aa} бүлэг", None))
        rows.append({
            "code": aa + "0000", "name": cmon, "name_en": cen,
            "type": class_type[aa], "parent": None, "note": None,
        })
    if aabb and aabb not in seen_group:
        seen_group.add(aabb)
        gmon, gen = group_names.get(aabb, (lf["name"], lf["name_en"]))
        rows.append({
            "code": aabb + "00", "name": gmon, "name_en": gen,
            "type": group_type[aabb], "parent": aa + "0000", "note": None,
        })
    rows.append({**lf, "parent": (aabb + "00") if aabb else (aa + "0000")})

# ── 5. id онооно, parent_id холбоно, SQL гаргана ─────────────────────────
id_by_code = {row["code"]: i + 1 for i, row in enumerate(rows)}


def q(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"


cols = [
    "id", "code", "name", "name_en", "type", "parent_id", "is_active",
    "note", "currency", "nature", "is_temp", "temp_percent",
]

out = []
out.append("-- " + "=" * 58)
out.append("-- accounts seed — Түмэн Тээх ХХК-ийн дансны төлөвлөгөө")
out.append("-- Эх сурвалж: Дансны_Төлөвлөгөө_TT.xlsx (151 навч + класс/бүлгийн эх данс)")
out.append("-- Код: AABBCC. Хоёр хэл (name/name_en), зардалд МУСГАА/татвар note.")
out.append("-- Хүснэгт хоосон үед л оруулна. Supabase SQL Editor-д ажиллуулна.")
out.append("-- Нийт: %d данс (%d навч + %d класс + %d бүлэг)."
          % (len(rows), len(leaves), len(seen_class), len(seen_group)))
out.append("-- " + "=" * 58)
out.append("")
out.append("DO $$")
out.append("BEGIN")
out.append("IF (SELECT COUNT(*) FROM accounts) = 0 THEN")
out.append("")
out.append("INSERT INTO accounts (%s) VALUES" % ", ".join(cols))

vals = []
for i, row in enumerate(rows):
    rec = [
        i + 1, row["code"], row["name"], row.get("name_en"), row["type"],
        id_by_code[row["parent"]] if row["parent"] else None, True,
        row.get("note"), "MNT", nature_of(row["type"]), False, 0,
    ]
    vals.append("  (" + ", ".join(q(x) for x in rec) + ")")
out.append(",\n".join(vals) + ";")
out.append("")
out.append("PERFORM setval(pg_get_serial_sequence('accounts','id'),")
out.append("               (SELECT MAX(id) FROM accounts));")
out.append("")
out.append("END IF;")
out.append("END $$;")
out.append("")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(out))

from collections import Counter
print("written", OUT)
print("нийт %d данс | навч %d | класс %d | бүлэг %d"
      % (len(rows), len(leaves), len(seen_class), len(seen_group)))
print("навч type:", dict(Counter(l["type"] for l in leaves)))
print("зардлын note-той:", len(exp_meta))
