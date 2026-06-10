# -*- coding: utf-8 -*-
"""
Стандарт IFRS загвар дансны төлөвлөгөө үүсгэгч (хоёр хэлтэй).

Гаралт:
  1) scripts/accounts-seed.sql                      — Supabase seed
  2) Desktop/2026/Дансны_Төлөвлөгөө_Стандарт.xlsx    — TT-тэй ижил форматтай Excel

Код бүтэц (6 оронтой):
  1xxxxx Хөрөнгө · 3xxxxx Өр төлбөр · 5xxxxx Өмч · 6xxxxx Орлого · 7-8xxxxx Зардал
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SQL_OUT = r"C:/finance 2.0/scripts/accounts-seed.sql"
XLSX_OUT = r"C:/Users/natsa/Desktop/2026/Дансны_Төлөвлөгөө_Стандарт.xlsx"

A, L, E, I, X = "asset", "liability", "equity", "income", "expense"
AK, PA = "Актив", "Пассив"

# (code, name_mon, name_en, type, parent_code)
COA = [
    # ── 1. ХӨРӨНГӨ / ASSETS ───────────────────────────────────────────────
    ("110000", "Мөнгөн хөрөнгө ба түүнтэй адилтгах", "Cash & Cash Equivalents", A, None),
    ("110100", "Кассын бэлэн мөнгө", "Petty Cash", A, "110000"),
    ("110200", "Харилцах дансны мөнгө", "Bank Current Account", A, "110000"),
    ("110300", "Хадгаламжийн дансны мөнгө", "Savings Account", A, "110000"),
    ("110400", "Гадаад валютын кассын мөнгө", "Foreign Currency Cash", A, "110000"),
    ("110500", "Гадаад валютын харилцах данс", "Foreign Currency Bank Account", A, "110000"),
    ("110600", "Карт болон цахим мөнгө", "Card & E-money", A, "110000"),
    ("110700", "Замд яваа мөнгө", "Cash in Transit", A, "110000"),
    ("110900", "Бусад мөнгөн хөрөнгө", "Other Cash", A, "110000"),

    ("120000", "Богино хугацаат хөрөнгө оруулалт", "Short-term Investments", A, None),
    ("120100", "Хугацаат хадгаламж (1 жил хүртэл)", "Term Deposit (< 1yr)", A, "120000"),
    ("120200", "Үнэт цаас (борлуулахад бэлэн)", "Marketable Securities", A, "120000"),
    ("120900", "Бусад богино хуг. хөрөнгө оруулалт", "Other Short-term Investments", A, "120000"),

    ("130000", "Дансны болон бусад авлага", "Trade & Other Receivables", A, None),
    ("130100", "Худалдан авагч, захиалагчийн авлага", "Trade Receivables", A, "130000"),
    ("130200", "Хүүгийн авлага", "Interest Receivable", A, "130000"),
    ("130300", "Ногдол ашгийн авлага", "Dividend Receivable", A, "130000"),
    ("130400", "Ажилтны авлага", "Employee Receivables", A, "130000"),
    ("130500", "Татварын авлага", "Tax Receivable", A, "130000"),
    ("130600", "НӨАТ-ын авлага", "VAT Receivable", A, "130000"),
    ("130700", "Холбоотой талуудын авлага", "Related Party Receivables", A, "130000"),
    ("130800", "Бусад авлага", "Other Receivables", A, "130000"),
    ("130900", "Эргэлзээтэй авлагын хасагдуулга (хасах)", "Allowance for Doubtful Debts", A, "130000"),

    ("140000", "Урьдчилж төлсөн тооцоо ба зардал", "Prepayments & Advances", A, None),
    ("140100", "Нийлүүлэгчид урьдчилж төлсөн", "Advances to Suppliers", A, "140000"),
    ("140200", "Урьдчилж төлсөн зардал", "Prepaid Expenses", A, "140000"),
    ("140300", "Урьдчилж төлсөн татвар", "Prepaid Tax", A, "140000"),
    ("140900", "Бусад эргэлтийн хөрөнгө", "Other Current Assets", A, "140000"),

    ("150000", "Бараа материал", "Inventory", A, None),
    ("150100", "Түүхий эд, үндсэн материал", "Raw Materials", A, "150000"),
    ("150200", "Туслах материал, сэлбэг", "Supplies & Spare Parts", A, "150000"),
    ("150300", "Дуусаагүй үйлдвэрлэл", "Work in Progress", A, "150000"),
    ("150400", "Бэлэн бүтээгдэхүүн", "Finished Goods", A, "150000"),
    ("150500", "Худалдааны бараа", "Merchandise", A, "150000"),
    ("150600", "Бага үнэтэй, түргэн элэгдэх зүйл", "Low-value Items", A, "150000"),
    ("150700", "Замд яваа бараа материал", "Inventory in Transit", A, "150000"),
    ("150900", "Бараа материалын үнэ бууралт (хасах)", "Inventory Write-down", A, "150000"),

    ("160000", "Үндсэн хөрөнгө", "Property, Plant & Equipment", A, None),
    ("160100", "Газар", "Land", A, "160000"),
    ("160200", "Барилга, байгууламж", "Buildings", A, "160000"),
    ("160300", "Машин, тоног төхөөрөмж", "Machinery & Equipment", A, "160000"),
    ("160400", "Тээврийн хэрэгсэл", "Vehicles", A, "160000"),
    ("160500", "Тавилга, эд хогшил", "Furniture & Fixtures", A, "160000"),
    ("160600", "Компьютер, оффисын тоног төхөөрөмж", "Computers & Office Equipment", A, "160000"),
    ("160700", "Барилгын явцад буй хөрөнгө", "Construction in Progress", A, "160000"),
    ("160800", "Бусад үндсэн хөрөнгө", "Other PP&E", A, "160000"),
    ("160900", "Хуримтлагдсан элэгдэл (хасах)", "Accumulated Depreciation", A, "160000"),

    ("170000", "Биет бус хөрөнгө", "Intangible Assets", A, None),
    ("170100", "Програм хангамж", "Software", A, "170000"),
    ("170200", "Лиценз, эрхийн зөвшөөрөл", "Licenses & Permits", A, "170000"),
    ("170300", "Барааны тэмдэг, патент", "Trademarks & Patents", A, "170000"),
    ("170400", "Гудвил", "Goodwill", A, "170000"),
    ("170800", "Бусад биет бус хөрөнгө", "Other Intangibles", A, "170000"),
    ("170900", "Хуримтлагдсан хорогдол (хасах)", "Accumulated Amortization", A, "170000"),

    ("180000", "Урт хугацаат хөрөнгө оруулалт ба бусад", "Long-term Investments & Other", A, None),
    ("180100", "Охин/хараат компанид оруулалт", "Investment in Subsidiaries", A, "180000"),
    ("180200", "Урт хугацаат хадгаламж", "Long-term Deposits", A, "180000"),
    ("180300", "Урт хугацаат авлага", "Long-term Receivables", A, "180000"),
    ("180400", "Хойшлогдсон татварын хөрөнгө", "Deferred Tax Asset", A, "180000"),
    ("180900", "Бусад урт хугацаат хөрөнгө", "Other Non-current Assets", A, "180000"),

    # ── 3. ӨР ТӨЛБӨР / LIABILITIES ────────────────────────────────────────
    ("310000", "Дансны болон бусад өглөг", "Trade & Other Payables", L, None),
    ("310100", "Нийлүүлэгч, гүйцэтгэгчийн өглөг", "Trade Payables", L, "310000"),
    ("310200", "Холбоотой талуудын өглөг", "Related Party Payables", L, "310000"),
    ("310300", "Урьдчилж авсан орлого", "Deferred Income", L, "310000"),
    ("310900", "Бусад өглөг", "Other Payables", L, "310000"),

    ("320000", "Цалин ба нийгмийн даатгалын өглөг", "Payroll & Social Insurance Payables", L, None),
    ("320100", "Цалингийн өглөг", "Wages Payable", L, "320000"),
    ("320200", "Нийгмийн даатгалын шимтгэлийн өглөг", "Social Insurance Payable", L, "320000"),
    ("320300", "ХХОАТ-ын суутгалын өглөг", "PIT Withholding Payable", L, "320000"),

    ("330000", "Татварын өр төлбөр", "Tax Payables", L, None),
    ("330100", "НӨАТ-ын өглөг", "VAT Payable", L, "330000"),
    ("330200", "ААНОАТ-ын өглөг", "Corporate Income Tax Payable", L, "330000"),
    ("330300", "Бусад татвар, хураамжийн өглөг", "Other Taxes Payable", L, "330000"),

    ("340000", "Богино хугацаат зээл ба хүүгийн өглөг", "Short-term Loans & Interest", L, None),
    ("340100", "Богино хугацаат зээл", "Short-term Loans", L, "340000"),
    ("340200", "Хүүгийн өглөг", "Interest Payable", L, "340000"),
    ("340300", "Ногдол ашгийн өглөг", "Dividend Payable", L, "340000"),

    ("350000", "Хуримтлагдсан өр төлбөр ба нөөц", "Accrued Liabilities & Provisions", L, None),
    ("350100", "Хуримтлагдсан зардлын өглөг", "Accrued Expenses", L, "350000"),
    ("350200", "Урьдчилгаа татвар, нөөц", "Provisions", L, "350000"),

    ("410000", "Урт хугацаат өр төлбөр", "Long-term Liabilities", L, None),
    ("410100", "Урт хугацаат зээл", "Long-term Loans", L, "410000"),
    ("410200", "Бонд, өрийн бичиг", "Bonds Payable", L, "410000"),
    ("410300", "Санхүүгийн түрээсийн өглөг", "Finance Lease Liability", L, "410000"),
    ("410400", "Хойшлогдсон татварын өр", "Deferred Tax Liability", L, "410000"),
    ("410900", "Бусад урт хугацаат өр төлбөр", "Other Non-current Liabilities", L, "410000"),

    # ── 5. ЭЗЭМШИГЧИЙН ӨМЧ / EQUITY ───────────────────────────────────────
    ("510000", "Эзэмшигчийн өмч", "Owner's Equity", E, None),
    ("510100", "Өмч (хувь нийлүүлсэн хөрөнгө)", "Share Capital", E, "510000"),
    ("510200", "Халаасны хувьцаа (хасах)", "Treasury Shares", E, "510000"),
    ("510300", "Нэмж төлөгдсөн капитал", "Additional Paid-in Capital", E, "510000"),

    ("520000", "Нөөц ба дахин үнэлгээ", "Reserves & Revaluation", E, None),
    ("520100", "Дахин үнэлгээний нөөц", "Revaluation Reserve", E, "520000"),
    ("520200", "Гадаад валютын хөрвүүлгийн нөөц", "FX Translation Reserve", E, "520000"),
    ("520300", "Заавал байгуулах нөөц", "Statutory Reserve", E, "520000"),

    ("530000", "Хуримтлагдсан ашиг (алдагдал)", "Retained Earnings", E, None),
    ("530100", "Тайлант үеийн цэвэр ашиг (алдагдал)", "Current Year Profit/(Loss)", E, "530000"),
    ("530200", "Өмнөх үеийн хуримтлагдсан ашиг", "Prior Year Retained Earnings", E, "530000"),

    # ── 6. ОРЛОГО / INCOME ────────────────────────────────────────────────
    ("610000", "Үндсэн үйл ажиллагааны орлого", "Operating Revenue", I, None),
    ("610100", "Бараа борлуулсны орлого", "Sales Revenue — Goods", I, "610000"),
    ("610200", "Ажил, үйлчилгээ үзүүлсний орлого", "Service Revenue", I, "610000"),
    ("610300", "Борлуулалтын хөнгөлөлт, буцаалт (хасах)", "Sales Discounts & Returns", I, "610000"),

    ("620000", "Бусад үйл ажиллагааны орлого", "Other Income", I, None),
    ("620100", "Түрээсийн орлого", "Rental Income", I, "620000"),
    ("620200", "Хүүгийн орлого", "Interest Income", I, "620000"),
    ("620300", "Ногдол ашгийн орлого", "Dividend Income", I, "620000"),
    ("620400", "Гадаад валютын ханшийн олз", "FX Gain", I, "620000"),
    ("620500", "Үндсэн хөрөнгө борлуулсны олз", "Gain on Disposal of PP&E", I, "620000"),
    ("620600", "Татаас, санхүүжилтийн орлого", "Grants & Subsidies", I, "620000"),
    ("620900", "Бусад орлого", "Miscellaneous Income", I, "620000"),

    # ── 7. ӨРТӨГ БА ЗАРДАЛ / COST & EXPENSE ──────────────────────────────
    ("710000", "Борлуулсан бараа (бүтээгдэхүүн)-ний өртөг", "Cost of Sales", X, None),
    ("710100", "Борлуулсан барааны өртөг", "Cost of Goods Sold", X, "710000"),
    ("710200", "Борлуулсан бүтээгдэхүүний өртөг", "Cost of Products Sold", X, "710000"),
    ("710300", "Үзүүлсэн ажил, үйлчилгээний өртөг", "Cost of Services", X, "710000"),

    ("720000", "Үйл ажиллагааны зардал", "Operating Expenses", X, None),
    ("720100", "Цалин хөлсний зардал", "Salaries & Wages", X, "720000"),
    ("720200", "Нийгмийн даатгалын шимтгэлийн зардал", "Social Insurance Expense", X, "720000"),
    ("720300", "Элэгдэл, хорогдлын зардал", "Depreciation & Amortization", X, "720000"),
    ("720400", "Түрээсийн зардал", "Rent Expense", X, "720000"),
    ("720500", "Шатахуун, тээврийн зардал", "Fuel & Transport", X, "720000"),
    ("720600", "Цахилгаан, дулаан, ус, холбоо", "Utilities & Communication", X, "720000"),
    ("720700", "Засвар үйлчилгээний зардал", "Repairs & Maintenance", X, "720000"),
    ("720800", "Бичиг хэрэг, оффисын зардал", "Office & Stationery", X, "720000"),
    ("720900", "Зар сурталчилгаа, маркетинг", "Advertising & Marketing", X, "720000"),
    ("721000", "Томилолт, зочломтгойн зардал", "Travel & Entertainment", X, "720000"),
    ("721100", "Хандив, тусламжийн зардал", "Donations", X, "720000"),
    ("721200", "Эргэлзээтэй авлагын зардал", "Bad Debt Expense", X, "720000"),
    ("721900", "Бусад үйл ажиллагааны зардал", "Other Operating Expenses", X, "720000"),

    ("810000", "Санхүүгийн зардал", "Finance Costs", X, None),
    ("810100", "Хүүгийн зардал", "Interest Expense", X, "810000"),
    ("810200", "Гадаад валютын ханшийн гарз", "FX Loss", X, "810000"),
    ("810300", "Банкны шимтгэл, хураамж", "Bank Charges", X, "810000"),

    ("820000", "Бусад зардал ба алдагдал", "Other Expenses & Losses", X, None),
    ("820100", "Үндсэн хөрөнгө хассаны гарз", "Loss on Disposal of PP&E", X, "820000"),
    ("820200", "Торгууль, алданги", "Penalties & Fines", X, "820000"),
    ("820900", "Бусад зардал", "Miscellaneous Expenses", X, "820000"),

    ("890000", "Татварын зардал", "Income Tax Expense", X, None),
    ("890100", "ААНОАТ-ын зардал", "Corporate Income Tax", X, "890000"),
    ("890200", "Хойшлогдсон татварын зардал", "Deferred Tax Expense", X, "890000"),
]

DEBIT_NORMAL = {A, X}
TYPE_ANGILAL = {A: "Хөрөнгө", L: "Өр төлбөр", E: "Өмч", I: "Орлого", X: "Зардал"}

# ── Санхүүгийн тайлангийн мөр (E-balance / СС №361 маягт) ─────────────────
# СБТ = Санхүүгийн байдлын тайлан (баланс), ОДТ = Орлогын дэлгэрэнгүй тайлан.
# Балансын класс (эхний 2 орон) -> тайлангийн мөр (яг E-balance код).
FS_BY_CLASS = {
    "11": "СБТ 1.1.1 Мөнгө, түүнтэй адилтгах хөрөнгө",
    "12": "СБТ 1.1.5 Бусад санхүүгийн хөрөнгө",
    "13": "СБТ 1.1.2 Дансны авлага",
    "14": "СБТ 1.1.7 Урьдчилж төлсөн зардал/тооцоо",
    "15": "СБТ 1.1.6 Бараа материал",
    "16": "СБТ 1.2.1 Үндсэн хөрөнгө",
    "17": "СБТ 1.2.2 Биет бус хөрөнгө",
    "18": "СБТ 1.2.4 Урт хугацаат хөрөнгө оруулалт",
    "31": "СБТ 2.1.1.1 Дансны өглөг",
    "32": "СБТ 2.1.1.2 Цалингийн өглөг",
    "33": "СБТ 2.1.1.3 Татварын өр",
    "34": "СБТ 2.1.1.5 Богино хугацаат зээл",
    "35": "СБТ 2.1.1.10 Бусад богино хугацаат өр төлбөр",
    "41": "СБТ 2.1.2.1 Урт хугацаат зээл",
    "51": "СБТ 2.3.1 Өмч",
    "52": "СБТ 2.3.4 Хөрөнгийн дахин үнэлгээний нэмэгдэл",
    "53": "СБТ 2.3.7 Хуримтлагдсан ашиг",
}
# Тусгай override — балансын нарийн мөр ба бүх орлого/зардлын данс.
FS_OVERRIDE = {
    # ── Баланс — нарийвчилсан мөр ──
    "130500": "СБТ 1.1.3 Татвар, НДШ-ийн авлага",
    "130600": "СБТ 1.1.3 Татвар, НДШ-ийн авлага",
    "130200": "СБТ 1.1.4 Бусад авлага",
    "130300": "СБТ 1.1.4 Бусад авлага",
    "130400": "СБТ 1.1.4 Бусад авлага",
    "130700": "СБТ 1.1.4 Бусад авлага",
    "130800": "СБТ 1.1.4 Бусад авлага",
    "130900": "СБТ 1.1.4 Бусад авлага",
    "140900": "СБТ 1.1.8 Бусад эргэлтийн хөрөнгө",
    "180400": "СБТ 1.2.6 Хойшлогдсон татварын хөрөнгө",
    "320200": "СБТ 2.1.1.4 НДШ-ийн өглөг",
    "340200": "СБТ 2.1.1.6 Хүүний өглөг",
    "340300": "СБТ 2.1.1.7 Ногдол ашгийн өглөг",
    "310300": "СБТ 2.1.1.8 Урьдчилж орсон орлого",
    "350200": "СБТ 2.1.1.9 Нөөц /өр төлбөр/",
    "410400": "СБТ 2.1.2.3 Хойшлогдсон татварын өр",
    "510200": "СБТ 2.3.2 Халаасны хувьцаа",
    "510300": "СБТ 2.3.3 Нэмж төлөгдсөн капитал",
    "520200": "СБТ 2.3.5 Гадаад валютын хөрвүүлэлтийн нөөц",
    "520300": "СБТ 2.3.6 Эздийн өмчийн бусад хэсэг",
    # ── Орлогын дэлгэрэнгүй тайлан ──
    "610100": "ОДТ 1 Борлуулалтын орлого (цэвэр)",
    "610200": "ОДТ 1 Борлуулалтын орлого (цэвэр)",
    "610300": "ОДТ 1 Борлуулалтын орлого (цэвэр)",
    "710100": "ОДТ 2 Борлуулсан бүтээгдэхүүний өртөг",
    "710200": "ОДТ 2 Борлуулсан бүтээгдэхүүний өртөг",
    "710300": "ОДТ 2 Борлуулсан бүтээгдэхүүний өртөг",
    "620100": "ОДТ 4 Түрээсийн орлого",
    "620200": "ОДТ 5 Хүүний орлого",
    "620300": "ОДТ 6 Ногдол ашгийн орлого",
    "620600": "ОДТ 8 Бусад орлого",
    "620900": "ОДТ 8 Бусад орлого",
    "720500": "ОДТ 9 Борлуулалт, маркетингийн зардал",
    "720900": "ОДТ 9 Борлуулалт, маркетингийн зардал",
    "720100": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720200": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720300": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720400": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720600": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720700": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "720800": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "721000": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "721100": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "721200": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "721900": "ОДТ 10 Ерөнхий ба удирдлагын зардал",
    "810100": "ОДТ 11 Санхүүгийн зардал",
    "810300": "ОДТ 11 Санхүүгийн зардал",
    "820200": "ОДТ 12 Бусад зардал",
    "820900": "ОДТ 12 Бусад зардал",
    "620400": "ОДТ 13 Гадаад валютын ханшийн зөрүүний олз (гарз)",
    "810200": "ОДТ 13 Гадаад валютын ханшийн зөрүүний олз (гарз)",
    "620500": "ОДТ 14 Үндсэн хөрөнгө данснаас хассаны олз (гарз)",
    "820100": "ОДТ 14 Үндсэн хөрөнгө данснаас хассаны олз (гарз)",
    "890100": "ОДТ 19 Орлогын татварын зардал",
    "890200": "ОДТ 19 Орлогын татварын зардал",
}


def fs_line_of(code, parent):
    if parent is None:
        return None  # эх данс (бүлэг) — тайлангийн мөргүй
    if code in FS_OVERRIDE:
        return FS_OVERRIDE[code]
    return FS_BY_CLASS.get(code[:2])


def nature_of(t):
    return AK if t in DEBIT_NORMAL else PA


def dt_kt(t):
    return "Дебет" if t in DEBIT_NORMAL else "Кредит"


# ── SQL гаргах ────────────────────────────────────────────────────────────
id_by_code = {r[0]: i + 1 for i, r in enumerate(COA)}


def q(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"


cols = [
    "id", "code", "name", "name_en", "type", "parent_id", "is_active",
    "note", "fs_line", "currency", "nature", "is_temp", "temp_percent",
]

out = [
    "-- " + "=" * 58,
    "-- accounts seed — Стандарт IFRS загвар дансны төлөвлөгөө (хоёр хэлтэй)",
    "-- Нийт: %d данс. Код: 1 хөрөнгө · 3 өр · 5 өмч · 6 орлого · 7-8 зардал." % len(COA),
    "-- Хүснэгт хоосон үед л оруулна. Supabase SQL Editor-д ажиллуулна.",
    "-- " + "=" * 58, "",
    "DO $$", "BEGIN", "IF (SELECT COUNT(*) FROM accounts) = 0 THEN", "",
    "INSERT INTO accounts (%s) VALUES" % ", ".join(cols),
]
vals = []
for i, (code, mon, en, t, parent) in enumerate(COA):
    rec = [i + 1, code, mon, en, t, id_by_code[parent] if parent else None,
           True, None, fs_line_of(code, parent), "MNT", nature_of(t), False, 0]
    vals.append("  (" + ", ".join(q(x) for x in rec) + ")")
out.append(",\n".join(vals) + ";")
out += ["", "PERFORM setval(pg_get_serial_sequence('accounts','id'),",
        "               (SELECT MAX(id) FROM accounts));", "",
        "END IF;", "END $$;", ""]
open(SQL_OUT, "w", encoding="utf-8").write("\n".join(out))

# ── Excel гаргах (TT-тэй ижил формат) ─────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Дансны Төлөвлөгөө"

hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
grp_fill = PatternFill("solid", fgColor="DDEBF7")
cls_fill = PatternFill("solid", fgColor="BDD7EE")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Дансны дугаар", "Дансны нэр (МОН)", "Дансны нэр (ENG)",
           "Ангилал", "Дт/Кт", "Тайлангийн мөр (СС №361)", "Тайлбар / Тэмдэглэл"]
ncol = len(headers)
ws.append(headers)
for c in range(1, ncol + 1):
    cell = ws.cell(1, c)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(vertical="center", horizontal="center")
    cell.border = border

for code, mon, en, t, parent in COA:
    is_class = parent is None
    row = [code, mon, en, TYPE_ANGILAL[t], "" if is_class else dt_kt(t),
           fs_line_of(code, parent) or "", ""]
    ws.append(row)
    r = ws.max_row
    for c in range(1, ncol + 1):
        ws.cell(r, c).border = border
    if is_class:
        for c in range(1, ncol + 1):
            ws.cell(r, c).fill = cls_fill
            ws.cell(r, c).font = Font(bold=True)
    else:
        ws.cell(r, 2).alignment = Alignment(indent=1)

ws.freeze_panes = "A2"
widths = [16, 42, 36, 12, 8, 34, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Хураангуй хуудас
ws2 = wb.create_sheet("Хураангуй")
ws2.append(["Стандарт IFRS загвар дансны төлөвлөгөө — Хураангуй"])
ws2["A1"].font = Font(bold=True, size=13)
ws2.append(["Код бүтэц: 6 оронтой — 1 Хөрөнгө · 3 Өр төлбөр · 5 Өмч · 6 Орлого · 7-8 Зардал"])
ws2.append([])
ws2.append(["Ангилал", "Нэр", "Бүлэг", "Дансны тоо"])
for c in range(1, 5):
    ws2.cell(4, c).fill = hdr_fill
    ws2.cell(4, c).font = hdr_font
from collections import defaultdict
cls_count = defaultdict(int)
cls_name = {}
for code, mon, en, t, parent in COA:
    cls = code[0]
    if parent is None:
        cls_name.setdefault(cls, [])
    cls_count[cls[:1]] += 1
# class-level summary by leading digit
# (эхлэх цифрүүд, label) — өр төлбөр 3 ба 4-ийг хамтад нь
order = [(("1", "2"), "1-2xxxx Хөрөнгө / Assets"),
         (("3", "4"), "3-4xxxx Өр төлбөр / Liabilities"),
         (("5",), "5xxxxx Эзэмшигчийн өмч / Equity"),
         (("6",), "6xxxxx Орлого / Income"),
         (("7", "8"), "7-8xxxx Өртөг ба зардал / Cost & Expense")]
total = 0
for digs, label in order:
    cnt = sum(1 for r in COA if r[0][0] in digs)
    groups = sum(1 for r in COA if r[0][0] in digs and r[4] is None)
    ws2.append(["·".join(digs) + "xxxx", label, groups, cnt])
    total += cnt
ws2.append(["НИЙТ", "Бүх данс", "", total])
ws2.cell(ws2.max_row, 1).font = Font(bold=True)
ws2.cell(ws2.max_row, 4).font = Font(bold=True)
for w, i in zip([12, 40, 8, 12], range(1, 5)):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb.save(XLSX_OUT)

from collections import Counter
print("SQL  ->", SQL_OUT)
print("XLSX ->", XLSX_OUT)
print("нийт %d данс |" % len(COA), dict(Counter(r[3] for r in COA)))
